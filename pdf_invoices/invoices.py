import os
from openpyxl import Workbook
import pdfplumber
import re
from datetime import datetime
import mysql.connector

# Função para inserir registros no banco de dados
def execute_insert(cursor, invoice_number, invoice_date, file_name, status):
    sql = "INSERT INTO invoice_records (invoice_number, invoice_date, file_name, status) VALUES (%s, %s, %s, %s)"
    val = (invoice_number, invoice_date, file_name, status)
    cursor.execute(sql, val)  # Executa o comando SQL de inserção com os valores fornecidos

# Função principal
def main():
    # INÍCIO DO PROCESSO

    # Conexão com o banco de dados
    db = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="process_invoices"
    )
    cursor = db.cursor()
    print("--- Conexão com o banco de dados estabelecida ---")

    # Obtém a lista de arquivos do diretório
    directory = 'pdf_invoices'  # Nome do diretório onde estão os arquivos PDF
    files = os.listdir(directory)  # Lista todos os arquivos no diretório
    files_quantity = len(files)  # Conta quantos arquivos foram encontrados

    if files_quantity == 0:
        raise Exception("Nenhum arquivo encontrado no diretório")

    # Criação do arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice Imports'

    # Define cabeçalhos no Excel
    ws['A1'] = 'Invoice #'
    ws['B1'] = 'Date'
    ws['C1'] = 'File Name'
    ws['D1'] = 'Status'

    # Encontra a primeira linha vazia na coluna D
    last_empty_line = 1
    while ws["D" + str(last_empty_line)].value is not None:
        last_empty_line += 1

    # PROCESSAMENTO DOS ARQUIVOS
    for file in files:
        try:
            # Abre o arquivo PDF e extrai o texto da primeira página
            with pdfplumber.open(directory + "/" + file) as pdf:
                first_page = pdf.pages[0]
                pdf_text = first_page.extract_text()

            # Padrões de expressão regular para encontrar o número e a data da fatura
            inv_number_re_pattern = r'INVOICE #(\d+)'  # Busca 'INVOICE #' seguido de números
            inv_date_re_pattern = r'DATE (\d{2}/\d{2}/\d{4})'  # Busca 'DATE' seguido por uma data no formato dd/mm/aaaa

            # Realiza a busca de número e data usando as expressões regulares
            match_number = re.search(inv_number_re_pattern, pdf_text)
            match_date = re.search(inv_date_re_pattern, pdf_text)

            # Se encontrar o número da fatura, insere no Excel, caso contrário lança uma exceção
            if match_number:
                ws['A{}'.format(last_empty_line)] = match_number.group(1)
            else:
                raise Exception("Número da fatura não encontrado")

            # Se encontrar a data da fatura, insere no Excel, caso contrário lança uma exceção
            if match_date:
                ws['B{}'.format(last_empty_line)] = match_date.group(1)
            else:
                raise Exception("Data da fatura não encontrada")

            # Preenche as outras colunas no Excel
            ws['C{}'.format(last_empty_line)] = file
            ws['D{}'.format(last_empty_line)] = "Completed"

            # Insere os dados no banco de dados
            execute_insert(cursor, match_number.group(1), match_date.group(1), file, "Completed")
            db.commit()  # Salva a inserção no banco de dados

            # Move para a próxima linha
            last_empty_line += 1

        # Tratamento de exceções durante o processamento de arquivos
        except Exception as e:
            print(f"Erro ao processar o arquivo: {e}")

            # Preenche o Excel com o status de erro e insere o erro no banco de dados
            ws['C{}'.format(last_empty_line)] = file
            ws['D{}'.format(last_empty_line)] = "Exceção: {}".format(e)

            execute_insert(cursor, "N/A", "N/A", file, "Exceção: {}".format(e))
            db.commit()

            # Move para a próxima linha
            last_empty_line += 1

    # Fecha a conexão com o banco de dados
    cursor.close()
    db.close()

    # Salva o arquivo Excel com o timestamp atual no nome
    full_now = str(datetime.now()).replace(":", "-")
    dot_index = full_now.index(".")
    now = full_now[:dot_index]
    wb.save("Invoices - {}.xlsx".format(now))

# Executa a função principal se o script for rodado diretamente
if __name__ == "__main__":
    main()